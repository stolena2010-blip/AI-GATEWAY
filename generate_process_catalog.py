"""
ניתוח תהליכי ייצור מ-796 קבצי SUMMARY — פירוק כל תהליך בנפרד עם ספירת מופעים.
Output: NEW FILES/CATALOG_processes.xlsx
"""

import re
import glob
import pandas as pd
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NEW_FILES_DIR = Path(__file__).parent / "NEW FILES"
OUTPUT = NEW_FILES_DIR / "CATALOG_processes_stats.xlsx"
EXCLUDE_FILE = Path(__file__).parent / "BOM" / "לא עושים.xlsx"


def load_excluded_processes() -> set[str]:
    """Load the exclusion list from לא עושים.xlsx."""
    if not EXCLUDE_FILE.exists():
        print("⚠️  לא נמצא קובץ 'לא עושים.xlsx' — לא מסנן תהליכים")
        return set()
    df = pd.read_excel(EXCLUDE_FILE)
    excluded = set()
    for _, row in df.iterrows():
        name = str(row.get("תהליך", "")).strip()
        if name and name != "nan":
            excluded.add(name)
    print(f"🚫 Loaded {len(excluded)} excluded processes from 'לא עושים.xlsx'")
    return excluded


# ── Classification rules ──────────────────────────────────────

MATERIAL_RE = re.compile(
    r'(?i)^(אלומיניום|פלדה|נירוסטה|פליז|נחושת|טיטניום|חומר|סגסוגת|'
    r'aluminum|steel|stainless|brass|copper|titanium|'
    r'al[- ]?\d|ss[- ]?\d|aisi|inconel|invar|kovar|maraging|'
    r'DUPLEX|ULTEM|G10|FR4|פוליאמיד|ניילון|CuNi|'
    r'\d{4}-[HT]|PH\d+[- ])'
)

COATING_KEYWORDS = [
    "אנודייז", "תמורה", "פסיבציה", "ציפוי", "ניקל", "כרום", "אבץ",
    "קדמיום", "בדיל", "כסף", "זהב", "נחושת אלקטרו",
    "אוקסיד", "טיפול חום", "חיסום", "הקשיה", "קדמה",
    "anodize", "anodise", "conversion", "passivat", "plat",
    "oxide", "heat treat",
]

PAINTING_KEYWORDS = [
    "צביעה", "פריימר", "צבע עליון", "טופקאוט", "לכה", "אפוקסי",
    "פוליאורתן", "paint", "primer", "topcoat", "lacquer", "epoxy",
    "איטום", "שכבת ביניים", "מערכת צבע",
]

COLOR_RE = re.compile(
    r'(?i)(RAL\s*\d{4}|FED[- ]STD|#\d{5}|pantone|'
    r'BLACK|WHITE|GREEN|RED|BLUE|YELLOW|GRAY|GREY|OLIVE|'
    r'שחור|לבן|ירוק|אדום|כחול|צהוב|אפור|חום|כתום|סגול|מט$|מבריק$)',
)

INSERT_RE = re.compile(
    r'(?i)(קשיחים|insert|helicoil|keensert|rivnut|pem|bollhoff|'
    r'loctite|standoff|bushing|self.clinch|402\d{6}|MS\d{5})',
)

AUX_KEYWORDS = [
    "מיסוך", "סימון", "חריטה", "הדפסת משי", "הזרקת דיו",
    "שימון", "התזת חול", "תזת חול", "ניקוי", "הכנת פני שטח", "בדיקת", "בדיקה",
    "שיחרור מימן", "שחרור מימן", "שחרור מתח", "הסרה", "הדבקה",
    "אריזה", "שבירת קצוות", "גימור", "תיוג", "TAG", "BAG",
    "VIBROPEN", "MARKER", "מרקר", "SERVICEABILITY", "מעכב קורוזיה",
    "masking", "marking", "engrav", "blast", "strip",
    "מילוי חריצ", "צריבה", "ריתוך", "השחז", "הסרת סיגים", "הסרת טאנג",
    "דיווח משקל", "חותמת", "הטבעה", "אפייה", "גסות פני",
    "הגנה על משטחים", "הדפסה", "אטימ", "צביעת טקסט",
    "התקנת מסמרות", "מסמרות",
    "הברגות", "משטחים מסומנים", "חורים", "עומק", "גובה אות",
    "רוחב קו", "אזור", "O-ring", "o-ring", "חריצי",
]


def classify_segment(text: str) -> str:
    """Classify a process segment into a category."""
    t = text.strip()
    if not t:
        return ""
    # Material — always first segment, skip
    if MATERIAL_RE.search(t):
        return "חומר גלם"
    # Inserts
    if INSERT_RE.search(t):
        return "קשיחים"
    # Pure color
    if COLOR_RE.search(t) and not any(kw in t for kw in ["ציפוי", "אנודייז", "צביעה", "פריימר"]):
        # Check if it's ONLY a color reference (not a process with color)
        stripped = COLOR_RE.sub("", t).strip(" ,|")
        if len(stripped) < 5:
            return "צבעים"
    # Coating
    for kw in COATING_KEYWORDS:
        if kw in t.lower() or kw in t:
            return "ציפוי"
    # Painting
    for kw in PAINTING_KEYWORDS:
        if kw in t.lower() or kw in t:
            return "צביעה"
    # Auxiliary
    for kw in AUX_KEYWORDS:
        if kw in t.lower() or kw in t:
            return "תהליכים מלווים"
    # Unknown
    return "אחר"


def normalize_process(text: str) -> str:
    """Clean and normalize a process name for counting."""
    t = text.strip()
    # Remove spec references for grouping (keep the core process name)
    # Remove "לפי XXX" trailing specs
    t = re.sub(r'\s+לפי\s+.+$', '', t)
    # Remove trailing RAFDOCS/MIL references
    t = re.sub(r'\s*(?:RAFDOCS|PS|MIL|AMS|ASTM|SAE)[- ][\w#()\- .~]+$', '', t, flags=re.IGNORECASE)
    # Remove measurements like "50+/-5 µm", "3µm", "12 מיקרון"
    t = re.sub(r'\s*\d+[+/\-±]*\d*\s*(?:µm|מיקרון|מ"מ|mm|inch)\s*(?:מינימום|minimum)?', '', t, flags=re.IGNORECASE)
    return t.strip().rstrip(',').strip()


def load_all_summaries() -> pd.DataFrame:
    """Load all SUMMARY files into one DataFrame."""
    pattern = str(NEW_FILES_DIR / "SUMMARY_all_results_*.xlsx")
    files = sorted(glob.glob(pattern))
    print(f"📂 Found {len(files)} SUMMARY files")

    dfs = []
    errors = 0
    for f in files:
        try:
            df = pd.read_excel(f)
            dfs.append(df)
        except Exception:
            errors += 1

    if not dfs:
        raise FileNotFoundError("No SUMMARY files found!")

    data = pd.concat(dfs, ignore_index=True)
    print(f"📊 Loaded {len(data)} drawings ({errors} read errors)")
    return data


def extract_processes(data: pd.DataFrame) -> dict[str, Counter]:
    """Parse all processes from the data, classify and count."""
    counters: dict[str, Counter] = {
        "ציפוי": Counter(),
        "צביעה": Counter(),
        "תהליכים מלווים": Counter(),
        "חומר גלם": Counter(),
        "צבעים": Counter(),
        "קשיחים": Counter(),
        "אחר": Counter(),
    }

    total_drawings = len(data)
    drawings_with_processes = 0

    for _, row in data.iterrows():
        # Prefer merged_processes (cleaner), fallback to process_summary_hebrew
        proc_text = ""
        for col in ["merged_processes", "process_summary_hebrew"]:
            val = row.get(col)
            if pd.notna(val) and str(val).strip():
                proc_text = str(val).strip()
                break

        if not proc_text:
            continue

        drawings_with_processes += 1

        # Split by pipe — each pipe segment is a major category
        segments = [s.strip() for s in proc_text.split("|") if s.strip()]

        for i, segment in enumerate(segments):
            # First segment is almost always material
            if i == 0 and MATERIAL_RE.search(segment):
                mat = normalize_process(segment)
                if mat:
                    counters["חומר גלם"][mat] += 1
                continue

            # Some segments contain comma-separated sub-processes
            # (e.g. "מיסוך הברגות, חריטה S/N, סימון תיוג")
            # But coating/painting descriptions can also have commas within spec
            category = classify_segment(segment)

            if category == "קשיחים":
                # Count inserts as a whole
                counters["קשיחים"][normalize_process(segment)] += 1
                continue

            if category == "צבעים":
                counters["צבעים"][normalize_process(segment)] += 1
                continue

            # For auxiliary segments with commas — split sub-processes
            if category == "תהליכים מלווים" and "," in segment:
                sub_parts = [s.strip() for s in segment.split(",") if s.strip()]
                for sp in sub_parts:
                    cat = classify_segment(sp)
                    norm = normalize_process(sp)
                    if norm and cat:
                        counters[cat][norm] += 1
            else:
                norm = normalize_process(segment)
                if norm and category:
                    counters[category][norm] += 1

    print(f"🔍 {drawings_with_processes}/{total_drawings} drawings have process data")
    return counters


def build_excel():
    data = load_all_summaries()
    counters = extract_processes(data)

    # ── Filter out excluded processes ──
    excluded = load_excluded_processes()
    if excluded:
        removed_total = 0
        for key in counters:
            to_remove = [name for name in counters[key] if name in excluded]
            for name in to_remove:
                removed_total += counters[key].pop(name)
        print(f"   Removed {removed_total} occurrences of excluded processes")

    sheet_config = [
        ("② ציפוי", "ציפוי", "E2EFDA"),
        ("③ צביעה", "צביעה", "DAEEF3"),
        ("④ תהליכים מלווים", "תהליכים מלווים", "FCE4D6"),
        ("① חומר גלם", "חומר גלם", "F2F2F2"),
        ("⑤ צבעים", "צבעים", "E4DFEC"),
        ("⑥ קשיחים", "קשיחים", "FFF2CC"),
        ("⑦ אחר", "אחר", "FFFFFF"),
    ]

    # Summary - category totals
    summary_rows = []
    total_all = 0
    for sheet_name, key, _ in sheet_config:
        c = counters[key]
        unique = len(c)
        total = sum(c.values())
        total_all += total
        summary_rows.append({
            "קטגוריה": sheet_name,
            "תהליכים ייחודיים": unique,
            "סה\"כ מופעים": total,
        })
    summary_rows.append({
        "קטגוריה": "סה\"כ",
        "תהליכים ייחודיים": sum(len(counters[k]) for _, k, _ in sheet_config),
        "סה\"כ מופעים": total_all,
    })
    df_summary = pd.DataFrame(summary_rows)

    # Combined sheet — all processes in one table: תהליך, סוג תהליך, מופעים
    combined_rows = []
    for sheet_name, key, _ in sheet_config:
        c = counters[key]
        for process_name, count in c.most_common():
            combined_rows.append({
                "תהליך": process_name,
                "סוג תהליך": key,
                "מופעים": count,
            })
    # Sort by count descending
    combined_rows.sort(key=lambda r: r["מופעים"], reverse=True)
    df_combined = pd.DataFrame(combined_rows)
    df_combined.index = range(1, len(df_combined) + 1)
    df_combined.index.name = "#"

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        df_combined.to_excel(writer, sheet_name="סיכום כל התהליכים")
        df_summary.to_excel(writer, sheet_name="סיכום לפי קטגוריה", index=False)
        for sheet_name, key, _ in sheet_config:
            c = counters[key]
            if not c:
                continue
            items = c.most_common()
            df = pd.DataFrame(items, columns=["תהליך", "מופעים"])
            df.index = range(1, len(df) + 1)
            df.index.name = "#"
            # Add percentage column
            total = sum(c.values())
            df["אחוז"] = df["מופעים"].apply(lambda x: f"{x/total*100:.1f}%")
            df.to_excel(writer, sheet_name=sheet_name[:31])

    # ── Format ──
    wb = load_workbook(OUTPUT)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell_font = Font(size=11, name="Calibri")
    top3_font = Font(size=11, name="Calibri", bold=True, color="C00000")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    all_sheets = [("סיכום כל התהליכים", "D6E4F0"), ("סיכום לפי קטגוריה", "D9E2F3")] + [(s[:31], bg) for s, _, bg in sheet_config]

    for sheet_name, bg_hex in all_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        data_fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.font = cell_font
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # Header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Alternating rows + bold top 3
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
            if idx % 2 == 0:
                for cell in row:
                    cell.fill = data_fill
            # Top 3 rows in red bold (skip summary sheet)
            if sheet_name != "סיכום" and idx <= 3:
                for cell in row:
                    cell.font = top3_font

        # Auto-width
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Row height
        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 22

        ws.sheet_view.rightToLeft = True

    wb.save(OUTPUT)
    print(f"\n✅ Saved: {OUTPUT}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")
    for sheet_name, key, _ in sheet_config:
        c = counters[key]
        if c:
            top3 = c.most_common(3)
            top_str = ", ".join(f"{name}({count})" for name, count in top3)
            print(f"   {sheet_name}: {len(c)} unique — Top: {top_str}")


if __name__ == "__main__":
    build_excel()
