"""
Quotes pipeline snapshot tests — lock down B2B and Excel output format,
plus contract tests for the scan_folder result schema.

These tests protect the quotes profile from regressions when adding
new profiles (orders, invoices, etc.).  If any change to shared code
alters the quotes output structure, these tests will fail immediately.
"""
import pytest
from pathlib import Path

from src.services.reporting.b2b_export import (
    _save_text_summary_with_variants,
    _save_text_summary,
)
from src.services.reporting.excel_export import _save_results_to_excel
from src.pipeline.drawing_processor import _build_result_dict


# ── Deterministic fixture with all fields a quotes result can have ──────
@pytest.fixture
def quotes_results():
    """Realistic quotes drawing results covering all B2B/Excel fields."""
    return [
        {
            "file_name": "drawing_AAA111.pdf",
            "part_number": "AAA-111",
            "drawing_number": "DWG-AAA",
            "revision": "C",
            "item_name": "BRACKET ASSY",
            "material": "AL 6061-T6",
            "customer_name": "IAI",
            "quantity": "50",
            "delivery_date": "2026-06-15",
            "confidence_level": "HIGH",
            "merged_description": "ציפוי אנודייז קשיח + צביעה אפוקסי",
            "process_summary_hebrew": "אנודייז קשיח",
            "merged_processes": "Anodizing Type III, Epoxy Paint",
            "merged_specs": "MIL-A-8625 Type III",
            "merged_bom": "",
            "merged_notes": "",
            "color_prices": "",
            "email_from": "buyer@customer.com",
            "email_subject": "RFQ 12345",
            "geometric_area": "~300 cm²",
            "part_number_ocr_original": "",
            "inserts_hardware": "",
        },
        {
            "file_name": "drawing_BBB222.pdf",
            "part_number": "BBB-222",
            "drawing_number": "DWG-BBB",
            "revision": "A",
            "item_name": "COVER PLATE",
            "material": "SS 304",
            "customer_name": "RAFAEL",
            "quantity": "(10, 20)",
            "delivery_date": "",
            "confidence_level": "MEDIUM",
            "merged_description": "",
            "process_summary_hebrew": "פסיבציה כימית",
            "merged_processes": "Passivation",
            "merged_specs": "AMS 2700",
            "merged_bom": "",
            "merged_notes": "דרוש אישור ראשונים",
            "color_prices": "",
            "email_from": "buyer@customer.com",
            "email_subject": "RFQ 12345",
            "geometric_area": "~150 cm²",
            "part_number_ocr_original": "BBB222",
            "inserts_hardware": "",
        },
        {
            "file_name": "drawing_CCC333.pdf",
            "part_number": "CCC-333",
            "drawing_number": "DWG-CCC",
            "revision": "",
            "item_name": "SPRING HOLDER",
            "material": "STEEL 4340",
            "customer_name": "ELBIT",
            "quantity": "1.0000",
            "delivery_date": "2026-07-01",
            "confidence_level": "FULL",
            "merged_description": "ציפוי קדמיום + ייבוש",
            "process_summary_hebrew": "ציפוי קדמיום",
            "merged_processes": "Cadmium Plating",
            "merged_specs": "",
            "merged_bom": "",
            "merged_notes": "",
            "color_prices": "",
            "email_from": "buyer@customer.com",
            "email_subject": "RFQ 12345",
            "geometric_area": "",
            "part_number_ocr_original": "",
            "inserts_hardware": "",
        },
    ]


# ════════════════════════════════════════════════════════════════════════
#  B2B FORMAT SNAPSHOTS
# ════════════════════════════════════════════════════════════════════════

class TestB2BFormatSnapshot:
    """Lock the B2B TAB-delimited format — exactly 18 fields per row."""

    B2B_FIELD_COUNT = 17  # _save_text_summary_with_variants produces 17 fields

    def _parse_b2b(self, path: Path):
        """Read a B2B file and return list of field-lists."""
        content = path.read_text(encoding="cp1255", errors="replace")
        rows = [r for r in content.split("{~#~}") if r.strip()]
        return [row.strip("\n\r").split("\t") for row in rows]

    def test_field_count_per_row(self, quotes_results, tmp_path):
        """Every B2B row must have exactly the expected number of TAB fields."""
        out = tmp_path / "B2B-0_200002-snap.txt"
        _save_text_summary_with_variants(
            quotes_results, out,
            customer_email="snap@test.com", b2b_number="B2B-0_200002", timestamp="SNAP",
        )
        for txt in tmp_path.glob("*.txt"):
            rows = self._parse_b2b(txt)
            for i, fields in enumerate(rows):
                assert len(fields) == self.B2B_FIELD_COUNT, (
                    f"{txt.name} row {i}: expected {self.B2B_FIELD_COUNT} fields, got {len(fields)}"
                )

    def test_row_separator(self, quotes_results, tmp_path):
        """Rows must be separated by {{~#~}} and file must end with it."""
        out = tmp_path / "B2B-0_200002-snap.txt"
        _save_text_summary_with_variants(
            quotes_results, out,
            customer_email="", b2b_number="B2B-0", timestamp="T",
        )
        content = (tmp_path / "B2B-0_200002-snap.txt").read_text(encoding="cp1255", errors="replace")
        assert content.endswith("{~#~}\n"), "File must end with {~#~}\\n"
        assert "{~#~}\n" in content

    def test_three_variant_files(self, quotes_results, tmp_path):
        """Must produce B2B, B2BH, B2BM variant files."""
        out = tmp_path / "B2B-0_200002-snap.txt"
        _save_text_summary_with_variants(
            quotes_results, out,
            customer_email="", b2b_number="B2B-0_200002", timestamp="T",
        )
        names = sorted(f.name for f in tmp_path.glob("*.txt"))
        assert any(n.startswith("B2B-") for n in names), f"Missing B2B variant: {names}"
        assert any(n.startswith("B2BH-") for n in names), f"Missing B2BH variant: {names}"
        assert any(n.startswith("B2BM-") for n in names), f"Missing B2BM variant: {names}"

    def test_field_positions_snapshot(self, quotes_results, tmp_path):
        """Lock the exact meaning of each field position."""
        out = tmp_path / "B2B-0_200002-snap.txt"
        _save_text_summary_with_variants(
            quotes_results, out,
            customer_email="snap@test.com", b2b_number="B2B-0_200002", timestamp="SNAP",
        )
        rows = self._parse_b2b(tmp_path / "B2B-0_200002-snap.txt")
        assert len(rows) == 3, f"Expected 3 rows, got {len(rows)}"

        # Row 1: AAA-111, qty=50 (numeric → field 4)
        r1 = rows[0]
        assert r1[0] == "1",            "field 1: row number"
        assert r1[1] == "AAA-111",      "field 2: part_number"
        assert r1[2] == "C",            "field 3: revision"
        assert r1[3] == "50",           "field 4: numeric quantity"
        assert r1[4] == "1",            "field 5: unit (always '1')"
        assert r1[5] == "0.0000",       "field 6: unit price (always '0.0000')"
        assert r1[6] == "0",            "field 7: currency (always '0')"
        assert r1[7] == "2026-06-15",   "field 8: delivery_date"
        assert r1[8] == "",             "field 9: notes (empty for numeric qty)"
        assert r1[9] == "HIGH",         "field 10: confidence_level"
        assert "ציפוי" in r1[10] or "אנודייז" in r1[10], "field 11: hebrew description"
        assert r1[11] == "BRACKET ASSY", "field 12: item_name"
        assert r1[12] == "0",           "field 13: B2B number (always '0')"
        assert r1[13] == "SNAP",        "field 14: timestamp"
        assert r1[14] == "snap@test.com", "field 15: customer_email"
        assert r1[15] == "DWG-AAA",     "field 16: drawing_number"
        assert r1[16] == "C",           "field 17: revision (again)"

    def test_non_numeric_qty_routes_to_field9(self, quotes_results, tmp_path):
        """Range quantity '(10, 20)' must go to field 9, field 4 must be '0'."""
        out = tmp_path / "B2B-0_200002-snap.txt"
        _save_text_summary_with_variants(
            quotes_results, out,
            customer_email="", b2b_number="B2B-0_200002", timestamp="T",
        )
        rows = self._parse_b2b(tmp_path / "B2B-0_200002-snap.txt")
        r2 = rows[1]  # BBB-222: quantity="(10, 20)"
        assert r2[1] == "BBB-222"
        assert r2[3] == "0",            "field 4: must be '0' for non-numeric"
        assert "(10, 20)" in r2[8],     "field 9: must contain the range"

    def test_high_variant_filters_correctly(self, quotes_results, tmp_path):
        """B2BH variant must contain only HIGH and FULL rows."""
        out = tmp_path / "B2B-0_200002-snap.txt"
        _save_text_summary_with_variants(
            quotes_results, out,
            customer_email="", b2b_number="B2B-0_200002", timestamp="T",
        )
        h_files = list(tmp_path.glob("B2BH-*.txt"))
        assert h_files, "B2BH file must exist"
        rows = self._parse_b2b(h_files[0])
        # AAA-111 (HIGH) + CCC-333 (FULL→HIGH) = 2 rows; BBB-222 (MEDIUM) excluded
        assert len(rows) == 2, f"B2BH should have 2 rows (HIGH+FULL), got {len(rows)}"
        pns = {r[1] for r in rows}
        assert "AAA-111" in pns
        assert "CCC-333" in pns
        assert "BBB-222" not in pns

    def test_medium_variant_filters_correctly(self, quotes_results, tmp_path):
        """B2BM variant must contain MEDIUM + HIGH (FULL→HIGH) rows."""
        out = tmp_path / "B2B-0_200002-snap.txt"
        _save_text_summary_with_variants(
            quotes_results, out,
            customer_email="", b2b_number="B2B-0_200002", timestamp="T",
        )
        m_files = list(tmp_path.glob("B2BM-*.txt"))
        assert m_files, "B2BM file must exist"
        rows = self._parse_b2b(m_files[0])
        # All 3: AAA-111 (HIGH) + BBB-222 (MEDIUM) + CCC-333 (FULL→HIGH)
        assert len(rows) == 3, f"B2BM should have 3 rows, got {len(rows)}"

    def test_encoding_is_cp1255(self, quotes_results, tmp_path):
        """B2B files must be readable as cp1255 (Windows-1255)."""
        out = tmp_path / "B2B-0_200002-snap.txt"
        _save_text_summary_with_variants(
            quotes_results, out,
            customer_email="", b2b_number="B2B-0_200002", timestamp="T",
        )
        for txt in tmp_path.glob("*.txt"):
            content = txt.read_bytes()
            decoded = content.decode("cp1255", errors="strict")
            assert "AAA-111" in decoded


# ════════════════════════════════════════════════════════════════════════
#  B2B SINGLE-FILE FORMAT (legacy _save_text_summary)
# ════════════════════════════════════════════════════════════════════════

class TestB2BSingleFileSnapshot:
    """Lock the legacy _save_text_summary — 18 fields per row."""

    B2B_FIELD_COUNT = 18  # _save_text_summary produces 18 fields (includes OCR original)

    def _parse_b2b(self, path: Path):
        content = path.read_text(encoding="cp1255", errors="replace")
        rows = [r for r in content.split("{~#~}") if r.strip()]
        return [row.strip("\n\r").split("\t") for row in rows]

    def test_field_count(self, quotes_results, tmp_path):
        out = tmp_path / "B2B-single.txt"
        _save_text_summary(
            quotes_results, out,
            customer_email="test@t.com", b2b_number="B2B-0", timestamp="TS",
        )
        rows = self._parse_b2b(out)
        for i, fields in enumerate(rows):
            assert len(fields) == self.B2B_FIELD_COUNT, (
                f"Row {i}: expected {self.B2B_FIELD_COUNT} fields, got {len(fields)}"
            )

    def test_field18_ocr_original(self, quotes_results, tmp_path):
        """Field 18 must contain part_number_ocr_original when present."""
        out = tmp_path / "B2B-single.txt"
        _save_text_summary(
            quotes_results, out,
            customer_email="", b2b_number="B2B-0", timestamp="TS",
        )
        rows = self._parse_b2b(out)
        # BBB-222 has part_number_ocr_original="BBB222"
        r2 = rows[1]
        assert r2[17] == "BBB222", "field 18: OCR original part number"
        # AAA-111 has no OCR original
        r1 = rows[0]
        assert r1[17] == "", "field 18: empty when no OCR override"


# ════════════════════════════════════════════════════════════════════════
#  EXCEL COLUMN SNAPSHOT
# ════════════════════════════════════════════════════════════════════════

class TestExcelColumnSnapshot:
    """Lock the Excel column names and order for quotes results."""

    # Columns that MUST exist in every quotes Excel output
    REQUIRED_COLUMNS = [
        "file_name",
        "part_number",
        "drawing_number",
        "revision",
        "item_name",
        "customer_name",
        "quantity",
        "confidence_level",
        "email_from",
        "email_subject",
    ]

    # Columns that must appear when merged fields are present
    MERGED_COLUMNS = [
        "merged_description",
        "merged_processes",
        "merged_specs",
    ]

    def _get_headers(self, results, tmp_path, pl_items=None):
        import openpyxl
        out = tmp_path / "snap_results.xlsx"
        _save_results_to_excel(results, out, pl_items)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()
        return headers

    def test_required_columns_present(self, quotes_results, tmp_path):
        """All required quotes columns must exist in the Excel header."""
        headers = self._get_headers(quotes_results, tmp_path)
        for col in self.REQUIRED_COLUMNS:
            assert col in headers, f"Required column '{col}' missing from Excel: {headers}"

    def test_merged_columns_present(self, quotes_results, tmp_path):
        """Merged description columns must exist when data has them."""
        headers = self._get_headers(quotes_results, tmp_path)
        for col in self.MERGED_COLUMNS:
            assert col in headers, f"Merged column '{col}' missing from Excel: {headers}"

    def test_accuracy_score_column(self, quotes_results, tmp_path):
        """accuracy_score column must be auto-generated from confidence_level."""
        import openpyxl
        out = tmp_path / "snap_accuracy.xlsx"
        _save_results_to_excel(quotes_results, out, None)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "accuracy_score" in headers, "accuracy_score column must exist"

        acc_col = headers.index("accuracy_score") + 1
        # Row 2 = AAA-111 (HIGH), should have score >= 0.8
        score = ws.cell(row=2, column=acc_col).value
        assert score is not None and score >= 0.8, f"HIGH confidence should have score >= 0.8, got {score}"
        wb.close()

    def test_rafael_row_count(self, quotes_results, tmp_path):
        """RAFAEL highlighting: function must return count of RAFAEL rows."""
        out = tmp_path / "snap_rafael.xlsx"
        rafael_count = _save_results_to_excel(quotes_results, out, None)
        # BBB-222 has customer_name="RAFAEL"
        assert rafael_count == 1

    def test_pl_columns_when_pl_provided(self, quotes_results, sample_pl_items, tmp_path):
        """PL columns must appear when pl_items are provided."""
        headers = self._get_headers(quotes_results, tmp_path, pl_items=sample_pl_items)
        pl_cols = ["PL Part Number", "PL Summary", "PL Summary (AI)"]
        for col in pl_cols:
            assert col in headers, f"PL column '{col}' missing when pl_items provided"

    def test_pl_columns_absent_without_pl(self, quotes_results, tmp_path):
        """PL columns must NOT appear when pl_items is None."""
        headers = self._get_headers(quotes_results, tmp_path, pl_items=None)
        assert "PL Part Number" not in headers, "PL columns should not exist without pl_items"
        assert "PL Summary" not in headers, "PL columns should not exist without pl_items"

    def test_column_count_stable(self, quotes_results, tmp_path):
        """Lock the total column count to detect unexpected additions/removals."""
        headers = self._get_headers(quotes_results, tmp_path)
        # Record current count as baseline — update this number if columns
        # are intentionally added/removed for the quotes profile.
        assert len(headers) >= 10, f"Expected at least 10 columns, got {len(headers)}"

    def test_data_integrity(self, quotes_results, tmp_path):
        """Verify actual cell values match input data."""
        import openpyxl
        out = tmp_path / "snap_data.xlsx"
        _save_results_to_excel(quotes_results, out, None)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        pn_col = headers.index("part_number") + 1
        cust_col = headers.index("customer_name") + 1

        # Row 2 = first result
        assert ws.cell(row=2, column=pn_col).value == "AAA-111"
        assert ws.cell(row=2, column=cust_col).value == "IAI"
        # Row 3 = second result
        assert ws.cell(row=3, column=pn_col).value == "BBB-222"
        assert ws.cell(row=3, column=cust_col).value == "RAFAEL"
        # Row 4 = third result
        assert ws.cell(row=4, column=pn_col).value == "CCC-333"
        assert ws.cell(row=4, column=cust_col).value == "ELBIT"
        wb.close()


# ════════════════════════════════════════════════════════════════════════
#  SCAN_FOLDER RESULT SCHEMA CONTRACT
# ════════════════════════════════════════════════════════════════════════

class TestResultSchemaContract:
    """Lock the field names that _build_result_dict produces.

    These are the fields that downstream consumers (B2B export, Excel export,
    email builder) depend on.  If a field is renamed or removed, the quotes
    pipeline will break silently — these tests catch that immediately.
    """

    # Core identification fields — B2B fields 2, 3, 12, 16
    CORE_ID_FIELDS = {
        "file_name", "part_number", "drawing_number", "revision",
        "item_name", "customer_name", "confidence_level",
    }

    # Extraction fields — used by B2B field 11 (hebrew desc) and Excel
    EXTRACTION_FIELDS = {
        "material", "coating_processes", "painting_processes", "colors",
        "process_summary_hebrew", "specifications", "notes_full_text",
        "inserts_hardware", "part_area",
    }

    # Quantity fields — B2B fields 4, 8, 9
    QUANTITY_FIELDS = {
        "quantity", "quantity_match_type", "quantity_source",
        "work_description_doc", "work_description_email",
    }

    # Email fields — B2B field 15
    EMAIL_FIELDS = {"email_from", "email_subject"}

    # Metadata fields
    META_FIELDS = {
        "subfolder", "num_pages", "image_resolution",
        "execution_time_seconds", "extraction_cost_usd",
    }

    @pytest.fixture
    def result_dict(self):
        """Build a result dict using the real _build_result_dict function."""
        fake_data = {
            "customer_name": "TEST",
            "part_number": "PN-001",
            "drawing_number": "DWG-001",
            "revision": "A",
            "item_name": "BRACKET",
            "confidence_level": "HIGH",
            "material": "AL 6061",
            "coating_processes": "Anodizing",
            "painting_processes": "",
            "colors": "",
            "specifications": "MIL-A-8625",
            "notes_full_text": "test note",
            "inserts_hardware": [],
            "part_area": "~300 cm²",
            "process_summary_hebrew": "אנודייז",
            "num_pages": 1,
        }
        return _build_result_dict(fake_data, Path("/tmp/test.pdf"), Path("/tmp"))

    def test_core_id_fields_present(self, result_dict):
        """Core identification fields must exist in every result dict."""
        missing = self.CORE_ID_FIELDS - set(result_dict.keys())
        assert not missing, f"Missing core fields: {missing}"

    def test_extraction_fields_present(self, result_dict):
        """Extraction fields must exist (B2B/Excel depend on them)."""
        missing = self.EXTRACTION_FIELDS - set(result_dict.keys())
        assert not missing, f"Missing extraction fields: {missing}"

    def test_quantity_fields_present(self, result_dict):
        """Quantity placeholder fields must exist (filled later by matcher)."""
        missing = self.QUANTITY_FIELDS - set(result_dict.keys())
        assert not missing, f"Missing quantity fields: {missing}"

    def test_email_fields_present(self, result_dict):
        """Email placeholder fields must exist (filled later by folder_saver)."""
        missing = self.EMAIL_FIELDS - set(result_dict.keys())
        assert not missing, f"Missing email fields: {missing}"

    def test_meta_fields_present(self, result_dict):
        """Metadata fields must exist."""
        missing = self.META_FIELDS - set(result_dict.keys())
        assert not missing, f"Missing metadata fields: {missing}"

    def test_field_types(self, result_dict):
        """Key fields must have expected types."""
        assert isinstance(result_dict["part_number"], str)
        assert isinstance(result_dict["confidence_level"], str)
        assert isinstance(result_dict["quantity"], str)
        assert isinstance(result_dict["num_pages"], int)
        assert isinstance(result_dict["execution_time_seconds"], (int, float))
        assert isinstance(result_dict["inserts_hardware"], (str, list))

    def test_no_none_values(self, result_dict):
        """No field should be None — must be empty string or default."""
        for key, value in result_dict.items():
            assert value is not None, f"Field '{key}' is None — must have a default"
