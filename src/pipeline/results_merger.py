"""
results_merger — Merge per-folder results into SUMMARY files in NEW FILES.
Extracted from customer_extractor_v3_dual.scan_folder() post-loop merge.
"""
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from src.services.file.file_utils import _build_drawing_part_map
from src.utils.logger import get_logger

logger = get_logger(__name__)


def merge_all_results(
    subfolders_to_process: List[Path],
    output_folder: Path,
    timestamp: str,
    all_results: List[Dict[str, Any]],
    all_classifications: List[Dict[str, Any]],
    *,
    global_classification_tokens_in: int = 0,
    global_classification_tokens_out: int = 0,
) -> Tuple[Optional[Path], Optional[Path], List[Dict[str, Any]]]:
    """Merge per-folder files into SUMMARY Excel files.

    Returns:
        (final_results_path, final_classification_path, all_file_classifications)
    """
    from src.services.reporting.excel_export import (
        _save_classification_report,
        _save_results_to_excel,
        _update_pl_sheet_with_associated_items,
    )
    from src.services.reporting.b2b_export import _save_text_summary_with_variants

    logger.info("=" * 70)
    logger.info("Merging files from all folders...")
    logger.info("=" * 70)

    # ── 1. Merge drawing_results ──
    logger.info("Merging drawing_results files...")
    all_drawing_results: List[Dict] = []
    all_pl_items: List[Dict] = []
    all_file_classifications: List[Dict] = []
    final_classification_path: Optional[Path] = None

    for target_folder in subfolders_to_process:
        for df_file in target_folder.glob("drawing_results_*.xlsx"):
            try:
                df = pd.read_excel(df_file, sheet_name='Sheet1')
                results = df.to_dict('records')
                all_drawing_results.extend(results)
                logger.info(f"Read {len(results)} rows from {df_file.name}")

                try:
                    df_pl = pd.read_excel(df_file, sheet_name='Parts_List_Items')
                    if not df_pl.empty:
                        pl_items = df_pl.to_dict('records')
                        normalized_items = []
                        for item in pl_items:
                            normalized_items.append({
                                'item_number': item.get('Item Number') or item.get('item_number'),
                                'description': item.get('Description') or item.get('description'),
                                'quantity': item.get('Quantity') or item.get('quantity'),
                                'pl_filename': item.get('PL Filename') or item.get('pl_filename'),
                                'matched_item_name': item.get('Matched Item') or item.get('matched_item_name'),
                                'matched_drawing_part_number': item.get('Drawing Part Number') or item.get('matched_drawing_part_number'),
                            })
                        pl_items_filtered = [i for i in normalized_items if i.get('item_number')]
                        if pl_items_filtered:
                            all_pl_items.extend(pl_items_filtered)
                            logger.info(f"Read {len(pl_items_filtered)} PL items from {df_file.name}")
                except Exception as e:
                    logger.info(f"(No PL sheet in this file: {str(e)[:50]})")
            except Exception as e:
                logger.error(f"Failed to read {df_file.name}: {e}")

    final_results_path: Optional[Path] = None

    if all_drawing_results:
        final_results_path = output_folder / f"SUMMARY_all_results_{timestamp}.xlsx"
        _save_results_to_excel(all_drawing_results, final_results_path, None)

        # Add PL Items sheet from sub-folders
        _add_pl_sheet_to_summary(final_results_path, subfolders_to_process)

        # Post-process SUMMARY file
        if all_file_classifications and final_classification_path and final_results_path.exists():
            logger.info("Updating SUMMARY Parts_List_Items with associated_items...")
            _update_pl_sheet_with_associated_items(final_results_path, final_classification_path)

        logger.info(f"✅ Combined drawing analysis: {final_results_path.name} ({len(all_drawing_results)} drawings)")
        if all_pl_items:
            logger.info(f"Includes {len(all_pl_items)} combined PL items")

        # Summary text file
        _save_summary_text(subfolders_to_process, all_drawing_results, output_folder, timestamp)
    else:
        logger.warning("⚠️ No drawings to merge")

    # ── 2. Merge file classifications ──
    logger.info("Merging file_classification files...")

    for fc in all_classifications:
        if 'file_path' in fc and not isinstance(fc['file_path'], Path):
            fc['file_path'] = Path(fc['file_path'])
        all_file_classifications.append(fc)

    logger.info(f"Found {len(all_file_classifications)} file classifications from current run")

    for target_folder in subfolders_to_process:
        for cf_file in target_folder.glob("file_classification_*.xlsx"):
            try:
                df = pd.read_excel(cf_file)
                for _, row in df.iterrows():
                    file_path_str = str(row['file_path'])
                    if any(str(fc.get('file_path', '')) == file_path_str for fc in all_file_classifications):
                        continue
                    all_file_classifications.append({
                        'file_path': Path(file_path_str),
                        'file_type': row['file_type'],
                        'description': row['description'],
                        'quote_number': row.get('quote_number', ''),
                        'order_number': row.get('order_number', ''),
                        'associated_item': row.get('associated_item', ''),
                        'extension': row.get('extension', ''),
                    })
            except Exception as e:
                logger.error(f"Failed to read {cf_file.name}: {e}")

    if all_file_classifications:
        final_classification_path = output_folder / f"SUMMARY_all_classifications_{timestamp}.xlsx"
        all_drawing_map = _build_drawing_part_map(all_file_classifications, all_results)
        _save_classification_report(
            all_file_classifications, output_folder,
            global_classification_tokens_in, global_classification_tokens_out,
            custom_filename=f"SUMMARY_all_classifications_{timestamp}.xlsx",
            drawing_map=all_drawing_map,
            drawing_results=all_results,
        )
        logger.info(f"Combined file mapping: {final_classification_path.name} ({len(all_file_classifications)} files)")
    else:
        logger.info("No mapping files to merge")

    return final_results_path, final_classification_path, all_file_classifications


def copy_folders_to_tosend(
    subfolders_to_process: List[Path],
    tosend_folder: str,
    folder_classifications_map: Dict[Path, List[Dict]],
    confidence_level: str,
    all_results: List[Dict[str, Any]],
) -> None:
    """Copy all processed folders to TO_SEND."""
    from src.services.file.file_utils import _copy_folder_to_tosend

    logger.info("=" * 70)
    logger.info("Copying all processed folders to TO_SEND...")
    logger.info("=" * 70)
    for target_folder in subfolders_to_process:
        fc_list = folder_classifications_map.get(target_folder, None)
        _copy_folder_to_tosend(target_folder, Path(tosend_folder), fc_list, confidence_level, all_results)


def print_final_summary(
    all_results: List[Dict[str, Any]],
    all_classifications: List[Dict[str, Any]],
    final_results_path: Optional[Path],
    final_classification_path: Optional[Path],
    output_folder: Path,
    timestamp: str,
    cost_summary: Dict[str, Any],
    subfolders_to_process: List[Path],
    skipped_large_files: List[Dict],
    total_rafael_rows: int,
) -> None:
    """Print final summary statistics to logger."""
    logger.info("=" * 70)
    logger.info("The following files were saved successfully:")
    logger.info("=" * 70)
    if final_results_path and final_results_path.exists():
        logger.info(f"SUMMARY_all_results_{timestamp}.xlsx")
        logger.info(f"- {len(all_results)} drawings combined")
    if final_classification_path and final_classification_path.exists():
        logger.info(f"SUMMARY_all_classifications_{timestamp}.xlsx")
        logger.info(f"- {len(all_classifications)} files combined")
    logger.info("Additional files in each subfolder:")
    logger.info("- drawing_results_[subfolder].xlsx")
    logger.info("- file_classification_[subfolder].xlsx")
    logger.info("=" * 70)

    if not all_results or not final_results_path or not final_results_path.exists():
        return

    df = pd.DataFrame(all_results)
    execution_time = cost_summary.get('execution_time', 0)

    logger.info("=" * 70)
    logger.info(f"Final Summary Excel file: {final_results_path}")
    logger.info(f"Total files processed: {len(all_results)}")
    logger.info(f"Folders processed: {len(subfolders_to_process)}")
    if skipped_large_files:
        logger.info(f"Files skipped (too large): {len(skipped_large_files)}")
    if total_rafael_rows > 0:
        logger.info(f"Total RAFAEL rows highlighted: {total_rafael_rows}")

    non_empty = lambda col: sum(1 for v in df[col] if v and str(v).strip())

    logger.info("Field Extraction Statistics:")
    for field in ('customer_name', 'part_number', 'item_name', 'drawing_number', 'revision'):
        logger.info(f"- {field}: {non_empty(field)}/{len(all_results)}")

    full_conf = sum(1 for v in df['confidence_level'] if v == 'full')
    high_conf = sum(1 for v in df['confidence_level'] if v == 'high')
    medium_conf = sum(1 for v in df['confidence_level'] if v == 'medium')
    low_conf = sum(1 for v in df['confidence_level'] if v == 'low')
    logger.info("Confidence Levels:")
    logger.info(f"-  Full: {full_conf}  High: {high_conf}  Medium: {medium_conf}  Low: {low_conf}")

    needs_review = sum(1 for v in df['needs_review'] if v and "בעייתי" in str(v))
    needs_check = sum(1 for v in df['needs_review'] if v and "בדיקה" in str(v))
    if needs_review or needs_check:
        logger.info(f"Quality Alerts: Problematic={needs_review}, NeedsVerification={needs_check}")

    for field in ('material', 'coating_processes', 'painting_processes', 'colors',
                  'part_area', 'specifications', 'parts_list_page',
                  'process_summary_hebrew', 'notes_full_text'):
        logger.info(f"- {field}: {non_empty(field)}/{len(all_results)}")

    minutes = int(execution_time // 60)
    seconds = int(execution_time % 60)
    time_str = f"{minutes}m {seconds}s" if minutes > 0 else f"{seconds}s"
    logger.info(f"Total execution time: {time_str}")
    logger.info(f"Average per drawing: {cost_summary.get('avg_time_per_drawing', 0):.1f}s")
    logger.info(f"API Costs: ${cost_summary.get('total_cost', 0):.4f}")
    logger.info("=" * 70)
    logger.info(f"All files saved in: {output_folder}")


# ===== internal helpers =====================================================

def _add_pl_sheet_to_summary(final_results_path: Path, subfolders: List[Path]) -> None:
    """Copy PL Items rows from sub-folder drawing_results into the SUMMARY file."""
    logger.info("Adding PL Items to summary file from sub-folders...")
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        all_pl_rows: list = []
        for target_folder in subfolders:
            for df_file in target_folder.glob("drawing_results_*.xlsx"):
                try:
                    df_pl = pd.read_excel(df_file, sheet_name='Parts_List_Items')
                    if not df_pl.empty:
                        all_pl_rows.extend(df_pl.values.tolist())
                        logger.info(f"Copied {len(df_pl)} rows from {df_file.name}")
                except Exception:
                    pass

        if not all_pl_rows:
            return

        wb = load_workbook(final_results_path)
        if "Parts_List_Items" in wb.sheetnames:
            ws_pl = wb["Parts_List_Items"]
        else:
            ws_pl = wb.create_sheet(title="Parts_List_Items")
            headers = [
                'PL Filename', 'Item Number', 'Description', 'Associated Item',
                'Matched Item', 'Drawing Part Number', 'Quantity', 'Processes',
                'Specifications', 'Product Tree', 'Item Type',
            ]
            ws_pl.append(headers)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws_pl[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in all_pl_rows:
            ws_pl.append(row)

        col_widths = [20, 15, 25, 18, 20, 18, 12, 25, 30, 35, 15]
        for idx, width in enumerate(col_widths, 1):
            ws_pl.column_dimensions[chr(64 + idx)].width = width
            ws_pl.column_dimensions[chr(64 + idx)].alignment = Alignment(wrap_text=True)

        wb.save(final_results_path)
        wb.close()
        logger.info(f"Added PL Items sheet with {len(all_pl_rows)} rows to SUMMARY")
    except Exception as e:
        logger.error(f"Error adding PL sheet to SUMMARY: {e}")


def _save_summary_text(
    subfolders: List[Path],
    all_drawing_results: List[Dict],
    output_folder: Path,
    timestamp: str,
) -> None:
    """Create summary text file in output folder."""
    from src.services.reporting.b2b_export import _save_text_summary_with_variants

    logger.info("📄 Creating summary text file...")
    customer_email = ""
    for tf in subfolders:
        email_file = tf / "email.txt"
        if email_file.exists():
            try:
                with open(email_file, 'r', encoding='utf-8', errors='ignore') as f:
                    first_line = f.readline().strip()
                    if first_line and "@" in first_line:
                        customer_email = first_line.replace("כתובת שולח:", "").replace("From:", "").strip()
                        break
            except Exception:
                pass

    request_id = timestamp
    for item in all_drawing_results:
        if item.get('quote_number'):
            request_id = item['quote_number']
            break
        elif item.get('order_number'):
            request_id = item['order_number']
            break

    b2b_number = "B2B-0_200002"
    text_path = output_folder / f"{b2b_number}-{request_id}.txt"
    _save_text_summary_with_variants(all_drawing_results, text_path, customer_email, b2b_number, request_id)
