"""
Analyze customer part-number patterns from SUMMARY Excel files.
Outputs a report Excel with:
  - Sheet "raw": all customer + PN + drawing# + confidence rows
  - Sheet "patterns": detected regex patterns per customer with example PNs
"""

import re
import os
import sys
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SUMMARY_DIR = Path("NEW FILES")
OUTPUT_FILE = Path("reports/customer_pn_patterns.xlsx")

# ── Customer name normalization ────────────────────────────────
# Maps variant spellings to a canonical name.
CUSTOMER_ALIASES = {
    "elbit systems": "Elbit Systems",
    "elbit systems aerospace division": "Elbit Systems",
    "elbit systems ew and sigint - elisra ltd": "Elbit Systems Elisra",
    "elbit systems ew and sigint – elisra ltd": "Elbit Systems Elisra",
    "elbit systems electro-optic elop ltd.": "Elbit Systems Elop",
    "elbit systems electro-optics elop ltd.": "Elbit Systems Elop",
    "elbit advanced technology center": "Elbit Systems",
    "elbit systems land": "Elbit Systems Land",
    "elbit systems ltd.": "Elbit Systems",
    "electro-optics": "Elbit Systems Elop",
    "electro-optics elop ltd.": "Elbit Systems Elop",
    "iai": "IAI",
    "israel aircraft industries": "IAI",
    "rafael": "RAFAEL",
    "applied materials": "Applied Materials",
    "applied materials pdc": "Applied Materials",
    "kla": "KLA",
    "kla corporation": "KLA",
    "kla+": "KLA",
    "kla-tencor": "KLA",
    "novatec": "NOVATEC",
    "novatec motion systems": "NOVATEC",
    "novatec precision systems": "NOVATEC",
    "novatech motion systems": "NOVATEC",
    "atec": "NOVATEC",
    "atec ltd.": "NOVATEC",
    "atec motion systems": "NOVATEC",
    "watec innovision systems": "NOVATEC",
    "plasan": "PLASAN",
    "plasan reem": "PLASAN",
    "plasanreem": "PLASAN",
    "imco industries group": "IMCO",
    "imco defense division": "IMCO",
    "imco industries ltd.": "IMCO",
    "orbit": "Orbit",
    "nvidia": "NVIDIA",
    "nvidia corporation": "NVIDIA",
    "cabiran": "Cabiran",
    "b.a microwaves ltd": "B.A. Microwaves",
    "b.a. microwaves": "B.A. Microwaves",
    "b.a. microwaves ltd": "B.A. Microwaves",
    "third eye": "ThirdEye",
    "thirdeye": "ThirdEye",
    "thirdEye": "ThirdEye",
    "drs rada technologies": "DRS RADA",
    "ors rada technologies": "DRS RADA",
    "rada": "DRS RADA",
    "telkoor": "Telkoor",
    "telkoor power supplies ltd.": "Telkoor",
    "matan": "MATAN",
    "matan digital printers": "MATAN",
    "kratos | gmi eyal": "KRATOS",
    "kronos | gmi eyal": "KRATOS",
    "hewlett-packard company": "HP",
    "hp indigo ltd.": "HP",
    "san challenging engineering": "Challenging Engineering",
    "challenging engineering": "Challenging Engineering",
    "zivav technologies": "Zivav",
    "אבקו מעברי מתכת בע\"מ": "AVCO",
    "avco": "AVCO",
}


def normalize_customer(name):
    """Normalize customer name using alias table."""
    key = name.strip().lower()
    return CUSTOMER_ALIASES.get(key, name.strip())


def collect_data():
    """Read all SUMMARY Excel files and collect customer+PN pairs."""
    rows = []
    files = sorted(SUMMARY_DIR.glob("SUMMARY_all_results_*.xlsx"))
    total = len(files)
    print(f"Found {total} SUMMARY files")

    for i, fp in enumerate(files, 1):
        if i % 100 == 0:
            print(f"  reading {i}/{total} ...")
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            ws = wb.active
            header = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = list(row)
                    continue
                rec = dict(zip(header, row))
                customer = normalize_customer(str(rec.get("customer_name") or ""))
                pn = str(rec.get("part_number") or "").strip()
                dn = str(rec.get("drawing_number") or "").strip()
                conf = str(rec.get("confidence_level") or "").strip()
                fname = str(rec.get("file_name") or "").strip()
                if customer and pn:
                    rows.append({
                        "customer": customer,
                        "part_number": pn,
                        "drawing_number": dn,
                        "confidence": conf,
                        "file_name": fname,
                        "source_file": fp.name,
                    })
            wb.close()
        except Exception as e:
            print(f"  SKIP {fp.name}: {e}")
    print(f"Collected {len(rows)} rows from {total} files")
    return rows


# ── Pattern detection ──────────────────────────────────────────

def classify_token(ch):
    """Map a character to a class: D=digit, L=letter, S=separator."""
    if ch.isdigit():
        return "D"
    if ch.isalpha():
        return "L"
    return "S"


def to_skeleton(pn):
    """Convert a PN to a skeleton pattern, e.g. 'FC-44061-310000-001' -> 'LL-DDDDD-DDDDDD-DDD'."""
    return "".join(classify_token(c) if classify_token(c) != "S" else c for c in pn)


def skeleton_to_regex(skel):
    """Convert skeleton to a concise regex, e.g. 'LL-DDDDD-DDDDDD-DDD' -> '[A-Z]{2}-\\d{5}-\\d{6}-\\d{3}'."""
    parts = []
    i = 0
    while i < len(skel):
        ch = skel[i]
        if ch == "D":
            j = i
            while j < len(skel) and skel[j] == "D":
                j += 1
            n = j - i
            parts.append(f"\\d{{{n}}}" if n > 1 else "\\d")
            i = j
        elif ch == "L":
            j = i
            while j < len(skel) and skel[j] == "L":
                j += 1
            n = j - i
            parts.append(f"[A-Z]{{{n}}}" if n > 1 else "[A-Z]")
            i = j
        else:
            parts.append(re.escape(ch))
            i += 1
    return "".join(parts)


def detect_patterns(rows):
    """Group PNs by customer, find dominant skeleton patterns."""
    by_customer = defaultdict(list)
    for r in rows:
        by_customer[r["customer"]].append(r["part_number"])

    results = []
    for customer in sorted(by_customer):
        pns = by_customer[customer]
        unique_pns = sorted(set(pns))
        skel_counter = Counter()
        skel_examples = defaultdict(list)

        for pn in unique_pns:
            skel = to_skeleton(pn)
            skel_counter[skel] += 1
            if len(skel_examples[skel]) < 5:
                skel_examples[skel].append(pn)

        # Sort by frequency
        for skel, count in skel_counter.most_common():
            regex = skeleton_to_regex(skel)
            examples = skel_examples[skel]
            results.append({
                "customer": customer,
                "pattern_regex": regex,
                "skeleton": skel,
                "unique_pns": count,
                "total_appearances": sum(1 for p in pns if to_skeleton(p) == skel),
                "examples": " | ".join(examples),
            })

    return results


# ── Excel output ───────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_report(rows, patterns):
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()

    # ── Sheet: patterns ──
    ws_pat = wb.active
    ws_pat.title = "patterns"
    pat_headers = ["Customer", "Regex Pattern", "Skeleton", "Unique PNs", "Total Appearances", "Examples"]
    write_header(ws_pat, pat_headers)

    for i, p in enumerate(patterns, 2):
        ws_pat.cell(row=i, column=1, value=p["customer"])
        ws_pat.cell(row=i, column=2, value=p["pattern_regex"])
        ws_pat.cell(row=i, column=3, value=p["skeleton"])
        ws_pat.cell(row=i, column=4, value=p["unique_pns"])
        ws_pat.cell(row=i, column=5, value=p["total_appearances"])
        ws_pat.cell(row=i, column=6, value=p["examples"])

    # Auto-width
    for col in range(1, len(pat_headers) + 1):
        ws_pat.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30
    ws_pat.auto_filter.ref = ws_pat.dimensions

    # ── Sheet: raw ──
    ws_raw = wb.create_sheet("raw")
    raw_headers = ["Customer", "Part Number", "Drawing Number", "Confidence", "File Name", "Source"]
    write_header(ws_raw, raw_headers)

    # Sort by customer, then PN
    sorted_rows = sorted(rows, key=lambda r: (r["customer"], r["part_number"]))
    for i, r in enumerate(sorted_rows, 2):
        ws_raw.cell(row=i, column=1, value=r["customer"])
        ws_raw.cell(row=i, column=2, value=r["part_number"])
        ws_raw.cell(row=i, column=3, value=r["drawing_number"])
        c4 = ws_raw.cell(row=i, column=4, value=r["confidence"])
        ws_raw.cell(row=i, column=5, value=r["file_name"])
        ws_raw.cell(row=i, column=6, value=r["source_file"])
        # Color-code confidence
        conf = (r["confidence"] or "").lower()
        if conf in ("full", "high"):
            c4.fill = HIGH_FILL
        elif conf in ("low", "unknown", "none", ""):
            c4.fill = LOW_FILL

    for col in range(1, len(raw_headers) + 1):
        ws_raw.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    ws_raw.auto_filter.ref = ws_raw.dimensions

    # ── Sheet: summary ──
    ws_sum = wb.create_sheet("summary")
    sum_headers = ["Customer", "Total PNs", "Unique PNs", "High/Full %", "Top Pattern", "Top Pattern Count"]
    write_header(ws_sum, sum_headers)

    by_cust = defaultdict(list)
    for r in rows:
        by_cust[r["customer"]].append(r)

    row_idx = 2
    for cust in sorted(by_cust):
        cust_rows = by_cust[cust]
        total = len(cust_rows)
        unique = len(set(r["part_number"] for r in cust_rows))
        high_full = sum(1 for r in cust_rows if (r["confidence"] or "").lower() in ("full", "high"))
        pct = f"{high_full / total * 100:.0f}%" if total else "0%"

        # Find top pattern
        top_pat = next((p for p in patterns if p["customer"] == cust), None)
        top_regex = top_pat["pattern_regex"] if top_pat else ""
        top_count = top_pat["unique_pns"] if top_pat else 0

        ws_sum.cell(row=row_idx, column=1, value=cust)
        ws_sum.cell(row=row_idx, column=2, value=total)
        ws_sum.cell(row=row_idx, column=3, value=unique)
        ws_sum.cell(row=row_idx, column=4, value=pct)
        ws_sum.cell(row=row_idx, column=5, value=top_regex)
        ws_sum.cell(row=row_idx, column=6, value=top_count)
        row_idx += 1

    for col in range(1, len(sum_headers) + 1):
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    ws_sum.auto_filter.ref = ws_sum.dimensions

    wb.save(OUTPUT_FILE)
    print(f"\nReport saved: {OUTPUT_FILE}")


def main():
    rows = collect_data()
    if not rows:
        print("No data found!")
        sys.exit(1)

    patterns = detect_patterns(rows)

    # Print quick summary to console
    print(f"\n{'='*60}")
    print(f"{'Customer':<20} {'Unique PNs':>10} {'Top Pattern'}")
    print(f"{'='*60}")
    by_cust = defaultdict(set)
    for r in rows:
        by_cust[r["customer"]].add(r["part_number"])
    for cust in sorted(by_cust):
        top = next((p for p in patterns if p["customer"] == cust), None)
        regex = top["pattern_regex"][:40] if top else "?"
        print(f"{cust:<20} {len(by_cust[cust]):>10} {regex}")

    write_report(rows, patterns)


if __name__ == "__main__":
    main()
